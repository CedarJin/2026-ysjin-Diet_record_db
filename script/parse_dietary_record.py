"""
Parse handwritten dietary records from PDF files using GPT-4o Vision.
Converts PDFs to PIL Images using convert_pdf_to_jpg.py, then to base64 PNG.
"""

import os
from pathlib import Path
from typing import Dict, List, Optional, Union
from dotenv import load_dotenv
from openai import OpenAI
from PIL import Image
import base64
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from convert_pdf_to_jpg import PDFToPNGConverter

# Load environment variables
load_dotenv()

# Initialize OpenAI client
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# Initialize PDF converter with max dimension 3000
converter = PDFToPNGConverter(max_long_side=2000)



def pdf_to_images(pdf_path: Union[str, Path]) -> List[Image.Image]:
    """
    Convert PDF pages to PIL Image objects for vision processing.
    
    Args:
        pdf_path: Path to the PDF file
    
    Returns:
        List of PIL Image objects (one per page)
    
    Raises:
        FileNotFoundError: If PDF file doesn't exist
        RuntimeError: If conversion fails
    """
    pdf_path_obj = Path(pdf_path)
    if not pdf_path_obj.exists():
        raise FileNotFoundError(f"PDF file not found: {pdf_path_obj}")
    
    try:
        # Get all pages as PIL Images
        images = converter.get_multiple_pages_as_pil(str(pdf_path))
        return images
    except Exception as e:
        raise RuntimeError(f"Failed to convert PDF to images: {e}")


def image_to_base64_png(image: Image.Image) -> str:
    """
    Convert PIL Image to base64-encoded PNG string.
    
    Args:
        image: PIL Image object
    
    Returns:
        Base64 encoded PNG image string
    
    Raises:
        ValueError: If image conversion fails
    """
    buffered = BytesIO()
    
    # Convert to RGB if necessary (PNG supports RGBA, but we'll use RGB for consistency)
    if image.mode in ('RGBA', 'LA', 'P'):
        # Create white background for transparent images
        rgb_image = Image.new('RGB', image.size, (255, 255, 255))
        if image.mode == 'P':
            image = image.convert('RGBA')
        if image.mode in ('RGBA', 'LA'):
            rgb_image.paste(image, mask=image.split()[-1] if image.mode == 'RGBA' else None)
        image = rgb_image
    
    # Save as PNG format
    image.save(buffered, format="PNG", optimize=True)
    img_str = base64.b64encode(buffered.getvalue()).decode()
    return img_str


def _normalize_header(name: str) -> str:
    """Normalize Excel column header for matching."""
    if name is None:
        return ""
    return str(name).strip().lower().replace(" ", "_").replace("-", "_")


def load_metadata_from_xlsx(raw_dir: Union[str, Path], record_id: str) -> Optional[Dict[str, str]]:
    """
    Load metadata for a record from record_meta.xlsx in the Raw directory.

    Expected columns (case-insensitive): record_id, subject_id, data_type, date, timepoint.
    data_type: 'pat' = Patternized diet, 'hab' = Habitual Diet.

    Args:
        raw_dir: Path to record/Raw (contains record_meta.xlsx).
        record_id: Record ID to look up (e.g. R0001).

    Returns:
        Dict with keys: record_id, subject_id, data_type, date, formatted_date,
        day_of_week, timepoint (empty if missing). None if file or row not found.
    """
    path = Path(raw_dir) / "record_meta.xlsx"
    if not path.exists():
        return None

    header_to_field = {
        "record_id": "record_id",
        "record id": "record_id",
        "subject_id": "subject_id",
        "subject id": "subject_id",
        "data_type": "data_type",
        "data type": "data_type",
        "date": "date",
        "timepoint": "timepoint",
    }

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]] if wb.sheetnames else wb.active
        if ws.max_row < 2:
            wb.close()
            return None

        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        col_to_field: Dict[int, str] = {}
        for col_idx, cell in enumerate(header_row):
            norm = _normalize_header(cell.value)
            if norm in header_to_field:
                col_to_field[col_idx] = header_to_field[norm]

        record_id_col = None
        for idx, f in col_to_field.items():
            if f == "record_id":
                record_id_col = idx
                break
        if record_id_col is None:
            wb.close()
            return None

        target_id = str(record_id).strip()
        for row in ws.iter_rows(min_row=2):
            if len(row) <= record_id_col:
                continue
            cell_val = row[record_id_col].value
            if cell_val is None:
                continue
            if str(cell_val).strip() != target_id:
                continue

            def get_val(field: str, default: str = "") -> str:
                for idx, f in col_to_field.items():
                    if f == field and idx < len(row):
                        v = row[idx].value
                        return "" if v is None else str(v).strip()
                return default

            date_str = get_val("date")
            formatted_date = date_str
            day_of_week = "Unknown"
            if date_str:
                try:
                    # Try YYYY-MM-DD
                    if len(date_str) == 10 and date_str[4] == "-":
                        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                    elif len(date_str) == 8:
                        date_obj = datetime.strptime(date_str, "%Y%m%d")
                    else:
                        date_obj = datetime.strptime(date_str[:10], "%Y-%m-%d")
                    day_of_week = date_obj.strftime("%a")
                    formatted_date = date_obj.strftime("%Y-%m-%d")
                except Exception:
                    pass

            data_type_val = get_val("data_type").lower() or ""

            wb.close()
            return {
                "record_id": target_id,
                "subject_id": get_val("subject_id"),
                "data_type": data_type_val,
                "date": date_str,
                "record_date": date_str,
                "formatted_date": formatted_date,
                "day_of_week": day_of_week,
                "timepoint": get_val("timepoint"),
            }
        wb.close()
        return None
    except Exception:
        try:
            wb.close()
        except Exception:
            pass
        return None


def parse_filename(filename: str) -> Dict[str, str]:
    """
    Extract record ID from filename. Filename is only the record ID (e.g. R0001.pdf).

    Args:
        filename: PDF filename string (e.g. R0001.pdf)

    Returns:
        Dict with 'record_id' only (e.g. {'record_id': 'R0001'}).
    """
    base_name = Path(filename).stem
    return {"record_id": base_name or filename.replace(".pdf", "")}


def parse_dietary_record(pdf_path: Union[str, Path], output_dir: Union[str, Path]) -> bool:
    """
    Parse a dietary record PDF using GPT-4o Vision and save as HTML.
    
    Always uses vision processing: converts PDF to PIL Images using convert_pdf_to_jpg.py,
    then converts to base64 PNG for GPT-4o Vision API.
    
    Args:
        pdf_path: Path to the PDF file
        output_dir: Directory to save the parsed HTML file
    
    Returns:
        True if parsing and saving succeeded, False otherwise
    """
    pdf_path_obj = Path(pdf_path)
    output_dir_obj = Path(output_dir)

    # Record ID from filename only (e.g. R0001.pdf -> R0001)
    filename = pdf_path_obj.name
    record_id = parse_filename(filename)["record_id"]

    # Metadata from record_meta.xlsx (same directory as PDF)
    raw_dir = pdf_path_obj.parent
    file_info = load_metadata_from_xlsx(raw_dir, record_id)
    if file_info is None:
        file_info = {
            "record_id": record_id,
            "subject_id": "unknown",
            "data_type": "unknown",
            "date": "",
            "record_date": "",
            "formatted_date": "unknown",
            "day_of_week": "Unknown",
            "timepoint": "",
        }

    print(f"Processing: {filename}")
    print(f"  Record ID: {file_info['record_id']}")
    print(f"  Subject ID: {file_info['subject_id']}")
    print(f"  Data Type: {file_info['data_type']}")
    print(f"  Date: {file_info.get('formatted_date', file_info.get('date', ''))}")
    if file_info.get("timepoint"):
        print(f"  Timepoint: {file_info['timepoint']}")

    # Convert PDF to images using converter
    print("\nConverting PDF to images using converter...")
    try:
        images = pdf_to_images(pdf_path_obj)
        print(f"  Converted {len(images)} page(s) to PIL Images")
    except Exception as e:
        print(f"❌ Error converting PDF to images: {e}")
        return False
    
    # Parse using vision
    if images:
        print("   Using GPT-4o to read images...\n")
        return parse_with_vision(images, file_info, output_dir_obj)
    else:
        print("   Failed to convert PDF to images.\n")
        return False


def parse_with_vision(images: List[Image.Image], file_info: Dict[str, str], output_dir: Path) -> bool:
    """
    Parse PDF images using GPT-4o Vision API.
    
    Args:
        images: List of PIL Image objects (one per page)
        file_info: Dictionary with parsed file information
        output_dir: Directory to save the parsed HTML file
    
    Returns:
        True if parsing and saving succeeded, False otherwise
    """
    # Construct the prompt
    system_prompt = "You are a helpful assistant that parses handwritten dietary records into structured HTML format. Always output valid HTML that preserves the original structure and formatting."
    
    user_prompt = f"""Help me identify an English handwritten food record in HTML format. The handwritten content in each cell needs to be accurately recognized. The handwriting may be rather messy. You should refer to the filling requirements for each column and try to restore the original meaning based on the handwritten content.


HTML FORMATTING REQUIREMENTS:
1. Use <!DOCTYPE html> with <html lang="en">
2. Include <meta charset="UTF-8" />
3. Set title: "FOOD RECORD (Handwritten Transcription)"
4. Include CSS in <style> tag:
   table {{ border-collapse: collapse; width: 100%; }}
   th, td {{ border: 1px solid #000; padding: 6px; vertical-align: top; }}
   th {{ background: #f2f2f2; }}
5. Add <h2>FOOD RECORD</h2> heading
6. Use <table> with <thead> and <tbody> sections
7. Keep column headers with <br> tags for line breaks within header cells



Rule 01: Keep the exact column header wording AND instructions, preserving line breaks with <br>.
Rule 02: Time of Day must be formatted exactly as "%I:%M %p" (zero-padded hour, e.g., 07:10 AM, 01:10 PM).
Rule 03: Do NOT guess empty cells, leave them blank.
Rule 04: Keep original wording in cell contents; do not "correct" words or combine fields across columns.
Rule 05: If handwriting touches a row or cell border, assign it to the row whose writing baseline it aligns with.
Rule 06: Do NOT merge multiple handwritten rows into one row. Each handwritten row should remain as a separate table row.

Please parse this handwritten dietary record and output it as clean, well-structured HTML. Maintain the table structure if present, and format all dates and times according to the rules above."""
    
    # Prepare messages with images
    messages = [
        {"role": "system", "content": system_prompt},
        {
            "role": "user",
            "content": [
                {"type": "text", "text": user_prompt}
            ]
        }
    ]
    
    # Add images to the message
    for i, image in enumerate(images):
        base64_image = image_to_base64_png(image)
        messages[1]["content"].append({
            "type": "image_url",
            "image_url": {
                "url": f"data:image/png;base64,{base64_image}",
                "detail": "high"  # High detail for better handwriting recognition
            }
        })
        print(f"  Processed page {i+1}/{len(images)}")
    
    print("\nSending to GPT-4o for parsing...")
    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=messages,
            max_completion_tokens=4096,
            temperature=0.3
        )
        
        html_output = response.choices[0].message.content
        
    except Exception as e:
        print(f"❌ Error calling GPT-4o API: {e}")
        return False
    
    return save_html_output(html_output, file_info, output_dir)


def _inject_metadata_into_html(html: str, file_info: Dict[str, str]) -> str:
    """Replace or insert metadata section in HTML with values from file_info (from record_meta.xlsx)."""
    soup = BeautifulSoup(html, "html.parser")
    body = soup.find("body") or soup

    # Build metadata paragraphs: Record ID, Subject ID, Data Type, Date of Record, Day of Week, Timepoint
    labels_values = [
        ("Record ID", file_info.get("record_id", "")),
        ("Subject ID", file_info.get("subject_id", "")),
        ("Data Type", file_info.get("data_type", "")),
        ("Date of Record", file_info.get("formatted_date", file_info.get("date", ""))),
        ("Day of Week", file_info.get("day_of_week", "")),
    ]
    if file_info.get("timepoint"):
        labels_values.append(("Timepoint", file_info["timepoint"]))

    # Remove existing metadata <p> with these labels (so we replace with xlsx values)
    meta_labels = {lv[0].lower().replace(" ", "") for lv in labels_values}
    for p in list(body.find_all("p")):
        strong = p.find("strong")
        if not strong:
            continue
        label = strong.get_text(strip=True).replace(":", "").strip().lower().replace(" ", "")
        if label in meta_labels:
            p.decompose()

    # Insert metadata after <h2>
    anchor = body.find("h2")
    for label, value in labels_values:
        p = soup.new_tag("p")
        strong = soup.new_tag("strong")
        strong.string = f"{label}:"
        p.append(strong)
        p.append(f" {value}")
        if anchor:
            anchor.insert_after(p)
            anchor = p
        else:
            body.insert(0, p)

    return str(soup)


def save_html_output(html_output: str, file_info: Dict[str, str], output_dir: Path) -> bool:
    """
    Save HTML output to file. Injects metadata from file_info (record_meta.xlsx) into the HTML.

    Args:
        html_output: HTML content string from GPT
        file_info: Dictionary with record_id, subject_id, data_type, formatted_date, day_of_week, timepoint
        output_dir: Directory to save the HTML file

    Returns:
        True if saving succeeded, False otherwise
    """
    # Clean up the response (sometimes GPT wraps HTML in markdown code blocks)
    if "```html" in html_output:
        html_output = html_output.split("```html")[1].split("```")[0].strip()
    elif "```" in html_output:
        html_output = html_output.split("```")[1].split("```")[0].strip()

    # Inject metadata from record_meta.xlsx so saved HTML has correct values
    html_output = _inject_metadata_into_html(html_output, file_info)

    # Save HTML file
    output_filename = f"{file_info['record_id']}_OCR.html"
    output_path = output_dir / output_filename

    try:
        # Ensure output directory exists
        output_dir.mkdir(parents=True, exist_ok=True)

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html_output)

        print(f"✅ Successfully saved parsed HTML to: {output_path}\n")
        return True

    except Exception as e:
        print(f"❌ Error saving HTML file: {e}\n")
        return False


def main():
    """Main function to parse dietary records in record/Raw directory (demo: first 3 records)."""
    import time
    
    # Paths
    raw_dir = Path("record/Raw")
    ocr_dir = Path("record/OCR")
    
    # Find all PDF files
    all_pdf_files = sorted(raw_dir.glob("*.pdf"))
    
    if not all_pdf_files:
        print(f"❌ No PDF files found in {raw_dir}")
        return
    
    # Demo mode: process only first 6 records
    pdf_files = all_pdf_files[:6]
    
    print(f"📋 Demo mode: Processing first 6 of {len(all_pdf_files)} PDF file(s)")
    print("=" * 60)
    
    # Track results
    successful = []
    failed = []
    start_time = time.time()
    
    # Process each PDF file
    for idx, pdf_path in enumerate(pdf_files, 1):
        file_start_time = time.time()
        print(f"\n[{idx}/{len(pdf_files)}] Processing: {pdf_path.name}")
        print("-" * 60)
        
        # Parse the record
        success = parse_dietary_record(str(pdf_path), str(ocr_dir))
        
        file_elapsed = time.time() - file_start_time
        
        if success:
            file_info = parse_filename(pdf_path.name)
            successful.append(file_info['record_id'])
            print(f"✅ Successfully processed {file_info['record_id']} ({file_elapsed:.1f}s)")
        else:
            file_info = parse_filename(pdf_path.name)
            failed.append(file_info['record_id'])
            print(f"❌ Failed to process {file_info['record_id']} ({file_elapsed:.1f}s)")
    
    # Summary
    total_elapsed = time.time() - start_time
    print("\n" + "=" * 60)
    print("PROCESSING SUMMARY")
    print("=" * 60)
    print(f"Total files: {len(pdf_files)}")
    print(f"✅ Successful: {len(successful)}")
    if successful:
        print(f"   Records: {', '.join(successful)}")
    print(f"❌ Failed: {len(failed)}")
    if failed:
        print(f"   Records: {', '.join(failed)}")
    print(f"⏱️  Total time: {total_elapsed:.1f}s ({total_elapsed/60:.1f} min)")
    if len(successful) > 0:
        avg_time = total_elapsed / len(successful)
        print(f"📊 Average time per file: {avg_time:.1f}s")
    print("=" * 60)


if __name__ == "__main__":
    main()
